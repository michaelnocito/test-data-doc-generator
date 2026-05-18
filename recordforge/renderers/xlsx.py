"""XLSX renderer using openpyxl.

Behavior preserved exactly from v1: disclaimer row, bold header row with
green fill, auto-width columns capped at 28 chars.
"""

import secrets
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from recordforge.core.faker_utils import sanitize_filename
from recordforge.core.models import GeneratedDoc

DISCLAIMER = "FICTIONAL TEST DATA ONLY - generated for testing, demo, or training use."


def render(dataset: str, rows: list[dict], output_dir: Path) -> GeneratedDoc:
    """Render dataset rows to an XLSX file."""
    stem = sanitize_filename(f"{dataset}_{secrets.token_hex(3)}")
    path = output_dir / f"{stem}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = dataset[:31]
    ws.append([DISCLAIMER])

    headers = list(rows[0].keys()) if rows else ["note"]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    bold = Font(bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = bold
        cell.fill = fill

    for row in rows:
        ws.append([row.get(h) for h in headers])

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 28)

    wb.save(str(path))
    return GeneratedDoc(path=path, doc_type=dataset, format="xlsx")
